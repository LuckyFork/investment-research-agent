import openpyxl
from app.doc_pipeline.parsers.base import BaseParser, ParsedBlock


class ExcelParser(BaseParser):
    def parse(self, file_path: str) -> list[ParsedBlock]:
        blocks: list[ParsedBlock] = []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append(" | ".join(cells))
            if rows:
                blocks.append(
                    ParsedBlock(
                        type="table",
                        text="\n".join(rows),
                        section_title=sheet_name,
                    )
                )

        wb.close()
        return blocks
